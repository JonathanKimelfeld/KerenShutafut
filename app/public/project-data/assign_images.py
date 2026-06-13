#!/usr/bin/env python3
"""
Match Excel image_file_name values to WordPress pins by title,
then set each pin's featured image. Skips pins that already have one.
"""

import subprocess, json, os, re

try:
    import openpyxl
except ImportError:
    subprocess.run(['python3', '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

WP      = '/Applications/Local.app/Contents/Resources/extraResources/bin/wp-cli/posix/wp'
WP_PATH = '/Users/User/Local Sites/keren-shutafut-map/app/public'
XLSX    = '/Users/User/Downloads/project_table_updated.xlsx'

def wp(*args):
    return subprocess.run([WP, f'--path={WP_PATH}'] + list(args),
                          capture_output=True, text=True)

def norm_title(s):
    s = str(s).strip().lower()
    s = re.sub(r'["“”‘’\'״׳]', '', s)
    s = re.sub(r'[\-–—]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def norm_file(name):
    name = os.path.splitext(name)[0]
    name = name.lower()
    name = re.sub(r'[\-_\s]+', ' ', name)
    name = re.sub(r'\s*(scaled|\d+x\d+)\s*', '', name)
    # strip parenthesized numbers like "(1)" or "1"
    name = re.sub(r'\s*\(\d+\)\s*$', '', name)
    return name.strip()

def levenshtein(a, b):
    if a == b: return 0
    if not a: return len(b)
    if not b: return len(a)
    prev = list(range(len(b) + 1))
    for ca in a:
        curr = [prev[0] + 1]
        for j, cb in enumerate(b):
            curr.append(min(prev[j] + (0 if ca == cb else 1),
                            curr[j] + 1, prev[j + 1] + 1))
        prev = curr
    return prev[-1]

# ── Load WordPress pins ────────────────────────────────────────────────────────
r = wp('post', 'list', '--post_type=pin', '--post_status=publish',
       '--fields=ID,post_title', '--format=json')
wp_pins = json.loads(r.stdout)  # [{'ID': '833', 'post_title': '...'}]

# Build {norm_title → wp_id}
title_to_wpid = {}
for p in wp_pins:
    nt = norm_title(p['post_title'])
    title_to_wpid[nt] = p['ID']

# ── Load media library (original files only) ──────────────────────────────────
r2 = wp('eval', r'''
$atts = get_posts(array("post_type"=>"attachment","posts_per_page"=>-1,"post_status"=>"any"));
foreach ($atts as $a) {
    $f = get_post_meta($a->ID, "_wp_attached_file", true);
    echo $a->ID . "\t" . basename($f) . "\n";
}
''')
media = []
for line in r2.stdout.strip().split('\n'):
    if '\t' in line:
        att_id, fname = line.split('\t', 1)
        if re.search(r'-\d+x\d+\.(jpg|jpeg|png|gif|webp|avif)$', fname):
            continue
        media.append((att_id.strip(), fname.strip()))

# ── Get current thumbnails ────────────────────────────────────────────────────
r3 = wp('eval', r'''
$pins = get_posts(array("post_type"=>"pin","posts_per_page"=>-1,"post_status"=>"any"));
foreach ($pins as $p) {
    $t = get_post_thumbnail_id($p->ID);
    echo $p->ID . "\t" . ($t ?: "") . "\n";
}
''')
current_thumbs = {}
for line in r3.stdout.strip().split('\n'):
    if '\t' in line:
        pid, tid = line.split('\t', 1)
        current_thumbs[pid.strip()] = tid.strip()

# ── Load Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX)
ws = wb.active
headers  = [cell.value for cell in ws[2]]
title_col = headers.index('post_title')
img_col   = headers.index('image_file_name')

excel_rows = []
for row in ws.iter_rows(min_row=3, values_only=True):
    title = row[title_col]
    img   = row[img_col]
    if not title:
        continue
    excel_rows.append((str(title).strip(), str(img).strip() if img else ''))

print(f"WP pins:        {len(wp_pins)}")
print(f"Excel rows:     {len(excel_rows)}")
print(f"Media items:    {len(media)}\n")

# ── Match Excel title → WP pin ────────────────────────────────────────────────
TITLE_DIST = 5   # max edit distance for title match
FILE_DIST  = 3   # max edit distance for filename match

assigned   = 0
already    = 0
no_wp_pin  = []
no_img     = []
no_file    = []

for excel_title, img_name in excel_rows:
    nt = norm_title(excel_title)

    # Find WP pin by title
    wp_id = title_to_wpid.get(nt)
    if not wp_id:
        # fuzzy title match
        best_dist, best_id = 999, None
        for ntw, wid in title_to_wpid.items():
            d = levenshtein(nt, ntw)
            if d < best_dist:
                best_dist, best_id = d, wid
        if best_dist <= TITLE_DIST:
            wp_id = best_id
        else:
            no_wp_pin.append(excel_title)
            continue

    # Already has thumbnail
    if current_thumbs.get(str(wp_id)):
        already += 1
        continue

    if not img_name:
        no_img.append((excel_title, wp_id))
        continue

    # Find media item
    qf = norm_file(img_name)
    best_dist, best_att, best_fname = 999, None, None
    for att_id, fname in media:
        fn = norm_file(fname)
        if fn == qf:
            best_dist, best_att, best_fname = 0, att_id, fname
            break
        d = levenshtein(qf, fn)
        if d < best_dist:
            best_dist, best_att, best_fname = d, att_id, fname

    if best_dist > FILE_DIST:
        no_file.append((excel_title, wp_id, img_name, best_fname, best_dist))
        continue

    # Assign thumbnail
    r4 = wp('post', 'meta', 'update', str(wp_id), '_thumbnail_id', str(best_att))
    if r4.returncode in (0, 1):
        label = 'exact' if best_dist == 0 else f'dist={best_dist}'
        print(f"  ✅ [WP{wp_id}] {excel_title[:50]}")
        print(f"       {img_name}  →  {best_fname}  ({label})")
        current_thumbs[str(wp_id)] = str(best_att)
        assigned += 1
    else:
        print(f"  ❌ [WP{wp_id}] {excel_title[:50]}: {r4.stderr.strip()}")

print(f"\n{'='*60}")
print(f"✅ Assigned        : {assigned}")
print(f"⏭  Already had img : {already}")
print(f"❌ No matching file : {len(no_file)}")
print(f"⚠️  No image in Excel: {len(no_img)}")

if no_file:
    print("\nCould not match image file (file likely not uploaded):")
    for t, wid, img, best, d in no_file:
        print(f"  WP{wid} | {t[:45]} | excel='{img}' | closest='{best}' (dist={d})")

if no_img:
    print("\nNo image name in Excel:")
    for t, wid in no_img:
        print(f"  WP{wid} | {t[:55]}")

if no_wp_pin:
    print("\nExcel rows with no matching WP pin:")
    for t in no_wp_pin:
        print(f"  {t[:60]}")
print('='*60)
