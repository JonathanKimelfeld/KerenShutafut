#!/usr/bin/env python3
"""
Update existing WordPress pins from the updated Excel table.
Skips: location, image_file_name.
Run from any directory — paths are absolute.
"""

import subprocess
import json
import csv
import sys
import os

try:
    import openpyxl
except ImportError:
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────

WP      = '/Applications/Local.app/Contents/Resources/extraResources/bin/wp-cli/posix/wp'
WP_PATH = '/Users/User/Local Sites/keren-shutafut-map/app/public'
XLSX    = '/Users/User/Downloads/project_table_updated.xlsx'
MAPPING = os.path.join(WP_PATH, 'project-data/id_mapping.csv')

TAXONOMY_MAP = {
    'geographic_region': 'geographic_region',
    'recruitment_cycle': 'activity_cycle',
    'target_audience':   'target_audience',
    'fields':            'domains',
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def wp(*args, input_data=None):
    cmd = [WP, f'--path={WP_PATH}'] + list(args)
    result = subprocess.run(cmd, capture_output=True, text=True, input=input_data)
    return result

def load_term_cache(taxonomy):
    """Return a {name: term_id} dict for the given taxonomy."""
    r = wp('term', 'list', taxonomy, '--fields=term_id,name', '--format=json')
    if r.returncode != 0:
        return {}
    try:
        return {t['name']: str(t['term_id']) for t in json.loads(r.stdout)}
    except Exception:
        return {}

def get_or_create_term(taxonomy, name, cache):
    name = name.strip()
    if name in cache:
        return cache[name]
    r = wp('term', 'create', taxonomy, name, '--porcelain')
    if r.returncode == 0:
        tid = r.stdout.strip()
        cache[name] = tid
        return tid
    # Term may already exist under a slightly different key — try fetching
    r2 = wp('term', 'list', taxonomy, f'--name={name}', '--fields=term_id', '--format=json')
    try:
        items = json.loads(r2.stdout)
        if items:
            tid = str(items[0]['term_id'])
            cache[name] = tid
            return tid
    except Exception:
        pass
    return None

# ── Load ID mapping  (custom_id → wordpress_id) ───────────────────────────────

id_map = {}
with open(MAPPING, newline='', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        id_map[str(int(float(row['custom_id'])))] = row['wordpress_id']

# ── Load Excel ────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(XLSX)
ws = wb.active

# Row 2 contains the column headers
headers = [cell.value for cell in ws[2]]
data_rows = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[0] is None:
        continue
    data_rows.append(dict(zip(headers, row)))

print(f"Loaded {len(data_rows)} rows from Excel")
print(f"Loaded {len(id_map)} ID mappings\n")

# ── Pre-load taxonomy term caches ─────────────────────────────────────────────

term_caches = {tax: load_term_cache(tax) for tax in TAXONOMY_MAP.values()}

# ── Update loop ───────────────────────────────────────────────────────────────

success = 0
skipped = 0
errors  = []

for row in data_rows:
    raw_id = row.get('id')
    if raw_id is None:
        continue
    custom_id = str(int(float(raw_id)))

    if custom_id not in id_map:
        print(f"⚠️  ID {custom_id} not in id_mapping — skipping")
        skipped += 1
        continue

    wp_id = id_map[custom_id]
    title = (row.get('post_title') or '').strip()
    print(f"[{custom_id} → WP {wp_id}] {title[:50]}")

    try:
        # ── Post title + content ──────────────────────────────────────────────
        content = (row.get('post_content') or '').strip()
        r = wp('post', 'update', wp_id,
               f'--post_title={title}',
               f'--post_content={content}')
        if r.returncode != 0:
            raise RuntimeError(f"post update failed: {r.stderr.strip()}")

        # ── Meta fields ───────────────────────────────────────────────────────
        meta_fields = {
            'title_en':      row.get('post_name_en'),
            'operating_org': row.get('operating_org'),
            'project_link':  row.get('project_link'),
            'latitude':      row.get('latitude'),
            'longitude':     row.get('longitude'),
        }
        for key, value in meta_fields.items():
            if value is None or str(value).strip() == '':
                continue
            r = wp('post', 'meta', 'update', wp_id, key, str(value).strip())
            if r.returncode not in (0, 1):  # 1 = value unchanged (ok)
                print(f"  ⚠️  meta {key}: {r.stderr.strip()}")

        # ── Taxonomies ────────────────────────────────────────────────────────
        for col, taxonomy in TAXONOMY_MAP.items():
            raw = row.get(col)
            if not raw:
                continue
            term_names = [t.strip() for t in str(raw).split('|') if t.strip()]
            if not term_names:
                continue

            term_ids = []
            for name in term_names:
                tid = get_or_create_term(taxonomy, name, term_caches[taxonomy])
                if tid:
                    term_ids.append(tid)
                else:
                    print(f"  ⚠️  Could not get/create term '{name}' in {taxonomy}")

            if term_ids:
                r = wp('post', 'term', 'set', wp_id, taxonomy, *term_ids)
                if r.returncode != 0:
                    print(f"  ⚠️  term set {taxonomy}: {r.stderr.strip()}")

        print(f"  ✅ Done")
        success += 1

    except Exception as e:
        msg = f"ID {custom_id} (WP {wp_id}): {e}"
        errors.append(msg)
        print(f"  ❌ {msg}")

# ── Summary ───────────────────────────────────────────────────────────────────

print(f"\n{'='*60}")
print(f"✅ Updated : {success}")
print(f"⚠️  Skipped : {skipped}")
print(f"❌ Errors  : {len(errors)}")
if errors:
    print("\nErrors:")
    for e in errors:
        print(f"  • {e}")
print('='*60)
