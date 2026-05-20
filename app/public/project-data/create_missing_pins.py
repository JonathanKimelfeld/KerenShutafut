#!/usr/bin/env python3
"""
Creates missing pins directly on the live Hostinger server via SSH+WP-CLI.
Reads the Excel table, skips pins that already exist, creates the rest.
"""

import subprocess
import json
import csv
import sys
import os

try:
    import openpyxl
except ImportError:
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────

SSH     = ['ssh', '-p', '65002', 'u678169283@shutafut-map.org']
WP_PATH = '/home/u678169283/domains/shutafut-map.org/public_html'
XLSX    = '/Users/User/Downloads/project_table_updated.xlsx'
MAPPING = '/Users/User/Local Sites/keren-shutafut-map/app/public/project-data/id_mapping.csv'

MISSING_CUSTOM_IDS = {
    '41', '51', '55', '57', '66', '69', '75', '76', '77', '83',
    '89', '90', '93', '96', '97', '106', '109', '114', '116',
    '119', '123', '124', '127', '133'
}

TAXONOMY_MAP = {
    'geographic_region': 'geographic_region',
    'recruitment_cycle': 'activity_cycle',
    'target_audience':   'target_audience',
    'fields':            'domains',
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def wp(*args):
    cmd = SSH + ['cd ' + WP_PATH + ' && wp ' + ' '.join(
        a.replace("'", "'\\''") for a in args
    )]
    # Use a cleaner approach with proper quoting
    remote_cmd = 'cd ' + WP_PATH + ' && wp ' + ' '.join(
        "'" + str(a).replace("'", "'\\''") + "'" for a in args
    )
    result = subprocess.run(SSH + [remote_cmd], capture_output=True, text=True)
    return result

def load_term_cache(taxonomy):
    r = wp('term', 'list', taxonomy, '--fields=term_id,name', '--format=json')
    if r.returncode != 0:
        return {}
    try:
        return {t['name']: str(t['term_id']) for t in json.loads(r.stdout)}
    except Exception:
        return {}

def get_or_create_term(taxonomy, name, cache):
    name = name.strip()
    if name in cache:
        return cache[name]
    r = wp('term', 'create', taxonomy, name, '--porcelain')
    if r.returncode == 0:
        tid = r.stdout.strip()
        cache[name] = tid
        return tid
    r2 = wp('term', 'list', taxonomy, f'--name={name}', '--fields=term_id', '--format=json')
    try:
        items = json.loads(r2.stdout)
        if items:
            tid = str(items[0]['term_id'])
            cache[name] = tid
            return tid
    except Exception:
        pass
    return None

# ── Load Excel ────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(XLSX)
ws = wb.active
headers = [cell.value for cell in ws[2]]

missing_rows = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[0] is None:
        continue
    d = dict(zip(headers, row))
    if d.get('id') is None:
        continue
    cid = str(int(float(d['id'])))
    if cid in MISSING_CUSTOM_IDS:
        missing_rows.append((cid, d))

print(f"Creating {len(missing_rows)} missing pins on live site...\n")

# ── Pre-load taxonomy caches ──────────────────────────────────────────────────

term_caches = {tax: load_term_cache(tax) for tax in TAXONOMY_MAP.values()}

# ── Create loop ───────────────────────────────────────────────────────────────

success = 0
errors  = []
new_id_map = {}  # custom_id -> new wp_id

for custom_id, row in missing_rows:
    title   = (row.get('post_title') or '').strip()
    content = (row.get('post_content') or '').strip()
    print(f"[custom_id={custom_id}] Creating: {title[:55]}")

    try:
        # ── Create the post ───────────────────────────────────────────────────
        r = wp('post', 'create',
               '--post_type=pin',
               '--post_status=publish',
               f'--post_title={title}',
               f'--post_content={content}',
               '--porcelain')

        if r.returncode != 0:
            raise RuntimeError(f"post create failed: {r.stderr.strip()}")

        wp_id = r.stdout.strip()
        new_id_map[custom_id] = wp_id
        print(f"  Created WP post ID: {wp_id}")

        # ── Meta fields ───────────────────────────────────────────────────────
        meta_fields = {
            'title_en':         row.get('post_name_en'),
            'operating_org':    row.get('operating_org'),
            'project_link':     row.get('project_link'),
            'latitude':         row.get('latitude'),
            'longitude':        row.get('longitude'),
            'custom_project_id': custom_id,
        }
        for key, value in meta_fields.items():
            if value is None or str(value).strip() == '':
                continue
            r = wp('post', 'meta', 'update', wp_id, key, str(value).strip())
            if r.returncode not in (0, 1):
                print(f"  ⚠️  meta {key}: {r.stderr.strip()}")

        # ── Taxonomies ────────────────────────────────────────────────────────
        for col, taxonomy in TAXONOMY_MAP.items():
            raw = row.get(col)
            if not raw:
                continue
            term_names = [t.strip() for t in str(raw).split('|') if t.strip()]
            if not term_names:
                continue

            term_ids = []
            for name in term_names:
                tid = get_or_create_term(taxonomy, name, term_caches[taxonomy])
                if tid:
                    term_ids.append(tid)
                else:
                    print(f"  ⚠️  Could not get/create term '{name}' in {taxonomy}")

            if term_ids:
                r = wp('post', 'term', 'set', wp_id, taxonomy, *term_ids)
                if r.returncode != 0:
                    print(f"  ⚠️  term set {taxonomy}: {r.stderr.strip()}")

        print(f"  ✅ Done")
        success += 1

    except Exception as e:
        msg = f"custom_id {custom_id}: {e}"
        errors.append(msg)
        print(f"  ❌ {msg}")

# ── Summary ───────────────────────────────────────────────────────────────────

print(f"\n{'='*60}")
print(f"✅ Created : {success}")
print(f"❌ Errors  : {len(errors)}")
if errors:
    print("\nErrors:")
    for e in errors:
        print(f"  • {e}")
print(f"\nNew WP IDs assigned:")
for cid, wpid in new_id_map.items():
    print(f"  custom_id={cid} → WP {wpid}")
print('='*60)
